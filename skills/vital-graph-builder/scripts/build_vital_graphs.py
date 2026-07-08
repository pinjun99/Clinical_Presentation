#!/usr/bin/env python3
"""
Skill 2: Vital Graph Builder.

Reads the review-first vitals workbook produced by Skill 1, collects the rows the
user ticked (checkbox in column A) on each vital sheet, re-lists them horizontally
on the SAME sheet (right of the tick table), and draws a native Excel chart there.

Chart styling mirrors the two references:
- BP           -> vertical double-arrow lines between Systolic and Diastolic
                  built as custom error bars on the hidden Systolic series
                  (the user-verified recipe in excel_bp_arrow_chart_steps.md;
                  high-low lines never render in this Excel) + data table,
                  y-axis visible, min 50 — matches the WhatsApp BP image.
- Other vitals -> single blue (#4472C4, TDM theme accent1) line + circle markers
                  + data table + visible y-axis — matches TDM graph.xlsx.

openpyxl limitations are worked around by patching the saved .xlsx in place:
- openpyxl strips Excel's native checkbox feature (FeaturePropertyBag), which is
  why ticks previously degraded to TRUE/FALSE text -> re-injected after save so
  column A stays a real tick column.
- openpyxl emits bare-bones axis XML that Excel won't render labels for, and its
  high-low-lines element never renders -> both are patched into the chart XML
  directly, modelled on the axis/line XML Excel itself wrote in TDM graph.xlsx.

Usage:
    python build_vital_graphs.py [workbook.xlsx]
"""

import re
import shutil
import sys
import zipfile
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.data_source import NumDataSource, NumRef
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.chart.marker import Marker
from openpyxl.chart.plotarea import DataTable
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineEndProperties, LineProperties
from openpyxl.utils import get_column_letter

DEFAULT_WB = "outputs/vitals_extract/vitals_extract.xlsx"

# Tick-table layout produced by Skill 1.
FIRST_DATA_ROW = 4      # first data row on each vital sheet
GRAPH_COL = 1           # column A = tick checkbox
DATE_COL = 2
TIME_COL = 3

# Where the horizontal re-list + chart go on the same sheet.
BLOCK_START_COL = 8     # column H
LABEL_ROW = 1
CHART_ANCHOR = "H5"

# One colour for every line vital, matching TDM graph.xlsx (theme accent1).
LINE_COLOUR = "4472C4"
ARROW_COLOUR = "404040"  # BP double-arrow colour (dark grey/black like WhatsApp)

# Old separate graph sheets from earlier versions, removed on run.
LEGACY_GRAPH_SHEETS = [
    "BP Graph", "Heart Rate Graph", "Respiratory Graph", "SpO2 Graph", "Pain Score Graph",
]

# Per-vital config. Units go in the chart title (like the references); no y-axis title.
VITALS = [
    {"sheet": "BP", "title": "Blood Pressure (mmHg)",
     "value_cols": [("Systolic", 5), ("Diastolic", 6)], "bp": True, "y_min": 50},
    {"sheet": "Heart Rate", "title": "Heart Rate (bpm)",
     "value_cols": [("Pulse (bpm)", 4)], "bp": False, "y_min": None},
    {"sheet": "Respiratory", "title": "Respiratory Rate (breaths/min)",
     "value_cols": [("Respiratory (breaths/min)", 4)], "bp": False, "y_min": None},
    {"sheet": "SpO2", "title": "SpO2 (%)",
     "value_cols": [("SpO2 (%)", 4)], "bp": False, "y_min": None},
    {"sheet": "Pain Score", "title": "Pain Score (0-10)",
     "value_cols": [("Pain Score (0-10)", 4)], "bp": False, "y_min": None},
]

VITAL_SHEETS = [v["sheet"] for v in VITALS]


def is_ticked(value):
    if value is True:
        return True
    if isinstance(value, (int, float)) and value == 1:
        return True
    if isinstance(value, str):
        return value.strip().lower() in {"true", "yes", "y", "1", "x", "✓"}
    return False


def collect_ticked_rows(ws):
    rows = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        has_data = any(
            ws.cell(r, c).value not in (None, "")
            for c in (DATE_COL, TIME_COL, 4, 5, 6)
        )
        if not has_data:
            continue
        if is_ticked(ws.cell(r, GRAPH_COL).value):
            rows.append(r)
    return rows


def resolve_value(wb, ws, row, col):
    """Return the literal value of a cell, following a formula link back to the
    'Vitals Extract' sheet when present. Static values (not formulas) render the
    chart data table / axis reliably; formula-linked cells do not."""
    v = ws.cell(row, col).value
    if isinstance(v, str) and v.startswith("="):
        m = re.search(r"Vitals Extract'?!\$?([A-Z]+)\$?(\d+)", v)
        if m and "Vitals Extract" in wb.sheetnames:
            vc = column_index_from_string(m.group(1))
            vr = int(m.group(2))
            return wb["Vitals Extract"].cell(vr, vc).value
        return None
    return v


def clear_previous_block(ws):
    """Remove charts and any prior horizontal block so re-runs stay idempotent."""
    ws._charts = []
    for row in ws.iter_rows(min_col=BLOCK_START_COL, max_col=BLOCK_START_COL + 200,
                            min_row=1, max_row=8):
        for cell in row:
            cell.value = None


def build_vital(wb, ws, cfg):
    ticked = collect_ticked_rows(ws)
    clear_previous_block(ws)

    value_rows = list(range(LABEL_ROW + 1, LABEL_ROW + 1 + len(cfg["value_cols"])))
    ws.cell(LABEL_ROW, BLOCK_START_COL - 1, "Date / Time")
    for i, (label, _c) in enumerate(cfg["value_cols"]):
        ws.cell(value_rows[i], BLOCK_START_COL - 1, label)

    if not ticked:
        ws.cell(LABEL_ROW, BLOCK_START_COL, "No rows ticked - tick the checkboxes and re-run.")
        return 0

    for j, src_row in enumerate(ticked):
        col = BLOCK_START_COL + j
        date = resolve_value(wb, ws, src_row, DATE_COL)
        time = resolve_value(wb, ws, src_row, TIME_COL)
        label = f"{'' if date is None else date}\n{'' if time is None else time}"
        cell = ws.cell(LABEL_ROW, col, label)
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
        for i, (_lbl, src_col) in enumerate(cfg["value_cols"]):
            ws.cell(value_rows[i], col, resolve_value(wb, ws, src_row, src_col))

    helper_row = None
    if cfg["bp"]:
        # Arrow-length helper row (Systolic - Diastolic) feeding the custom
        # error bars, per excel_bp_arrow_chart_steps.md.
        helper_row = value_rows[-1] + 1
        ws.cell(helper_row, BLOCK_START_COL - 1, "Sys-Dia (arrow length)")
        for j in range(len(ticked)):
            col = BLOCK_START_COL + j
            s_v = ws.cell(value_rows[0], col).value
            d_v = ws.cell(value_rows[1], col).value
            if isinstance(s_v, (int, float)) and isinstance(d_v, (int, float)):
                ws.cell(helper_row, col, s_v - d_v)

    last_col = BLOCK_START_COL + len(ticked) - 1
    build_chart(ws, cfg, value_rows, last_col, helper_row)
    return len(ticked)


def build_chart(ws, cfg, value_rows, last_col, helper_row=None):
    chart = LineChart()
    chart.title = cfg["title"]
    chart.height = 9
    chart.width = 26
    if cfg["y_min"] is not None:
        chart.y_axis.scaling.min = cfg["y_min"]

    cats = Reference(ws, min_col=BLOCK_START_COL, max_col=last_col,
                     min_row=LABEL_ROW, max_row=LABEL_ROW)
    for i, (label, _c) in enumerate(cfg["value_cols"]):
        s = Series(
            Reference(ws, min_col=BLOCK_START_COL, max_col=last_col,
                      min_row=value_rows[i], max_row=value_rows[i]),
            title=label,
        )
        chart.series.append(s)
    chart.set_categories(cats)

    chart.plot_area.dTable = DataTable(
        showHorzBorder=True, showVertBorder=True, showOutline=True, showKeys=True)
    chart.legend = None

    if cfg["bp"]:
        # Hide both series (they still feed the data table), then draw the
        # WhatsApp-style vertical double arrows as custom error bars on the
        # Systolic series: minus direction spanning down to Diastolic, no end
        # cap, dark 1.5pt line with triangle arrowheads on both ends. This is
        # the user-verified recipe from excel_bp_arrow_chart_steps.md.
        for s in chart.series:
            s.graphicalProperties = GraphicalProperties()
            s.graphicalProperties.line = LineProperties(noFill=True)
            s.marker = Marker(symbol="none")
        col0 = get_column_letter(BLOCK_START_COL)
        col1 = get_column_letter(last_col)
        diff_ref = f"'{ws.title}'!${col0}${helper_row}:${col1}${helper_row}"
        arrow_line = LineProperties(
            solidFill=ARROW_COLOUR, w=19050,
            headEnd=LineEndProperties(type="triangle", w="med", len="med"),
            tailEnd=LineEndProperties(type="triangle", w="med", len="med"))
        chart.series[0].errBars = ErrorBars(
            errDir="y", errBarType="minus", errValType="cust", noEndCap=True,
            minus=NumDataSource(numRef=NumRef(f=diff_ref)),
            spPr=GraphicalProperties(ln=arrow_line))
    else:
        for s in chart.series:
            s.smooth = False
            line = LineProperties(solidFill=LINE_COLOUR, w=19050)
            s.graphicalProperties = GraphicalProperties(ln=line)
            s.graphicalProperties.line = line
            mk = Marker(symbol="circle", size=6)
            mk.graphicalProperties = GraphicalProperties(solidFill=LINE_COLOUR)
            mk.graphicalProperties.line = LineProperties(solidFill=LINE_COLOUR)
            s.marker = mk

    ws.add_chart(chart, CHART_ANCHOR)


# ---------------------------------------------------------------------------
# Post-save patching of the .xlsx package.
# ---------------------------------------------------------------------------

A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
C_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"

# Axis XML modelled on what Excel wrote in TDM graph.xlsx: visible axis labels
# (delete=0, tickLblPos=nextTo) coexisting with the chart data table.
CAT_AX = (
    '<catAx><axId val="{axid}"/><scaling><orientation val="minMax"/></scaling>'
    '<delete val="0"/><axPos val="b"/>'
    '<numFmt formatCode="General" sourceLinked="1"/>'
    '<majorTickMark val="out"/><minorTickMark val="none"/><tickLblPos val="nextTo"/>'
    '<spPr><a:noFill/><a:ln w="9525" cap="flat" cmpd="sng" algn="ctr">'
    '<a:solidFill><a:schemeClr val="tx1"><a:lumMod val="25000"/><a:lumOff val="75000"/></a:schemeClr></a:solidFill>'
    '<a:round/></a:ln></spPr>'
    '<txPr><a:bodyPr rot="0" spcFirstLastPara="1" vertOverflow="ellipsis" vert="horz" wrap="square" anchor="ctr" anchorCtr="1"/>'
    '<a:lstStyle/><a:p><a:pPr><a:defRPr sz="900"/></a:pPr><a:endParaRPr lang="en-US"/></a:p></txPr>'
    '<crossAx val="{cross}"/><crosses val="autoZero"/><auto val="1"/>'
    '<lblAlgn val="ctr"/><lblOffset val="100"/><noMultiLvlLbl val="0"/></catAx>'
)

VAL_AX = (
    '<valAx><axId val="{axid}"/><scaling><orientation val="minMax"/>{minmax}</scaling>'
    '<delete val="0"/><axPos val="l"/>'
    '<majorGridlines><spPr><a:ln w="9525" cap="flat" cmpd="sng" algn="ctr">'
    '<a:solidFill><a:schemeClr val="tx1"><a:lumMod val="15000"/><a:lumOff val="85000"/></a:schemeClr></a:solidFill>'
    '<a:round/></a:ln></spPr></majorGridlines>'
    '<numFmt formatCode="General" sourceLinked="1"/>'
    '<majorTickMark val="none"/><minorTickMark val="none"/><tickLblPos val="nextTo"/>'
    '<spPr><a:noFill/><a:ln><a:noFill/></a:ln></spPr>'
    '<txPr><a:bodyPr rot="-60000000" spcFirstLastPara="1" vertOverflow="ellipsis" vert="horz" wrap="square" anchor="ctr" anchorCtr="1"/>'
    '<a:lstStyle/><a:p><a:pPr><a:defRPr sz="900"/></a:pPr><a:endParaRPr lang="en-US"/></a:p></txPr>'
    '<crossAx val="{cross}"/><crosses val="autoZero"/><crossBetween val="between"/></valAx>'
)

# Excel native checkbox feature (verbatim from the Skill 1 workbook before
# openpyxl stripped it).
FPB_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<FeaturePropertyBags xmlns="http://schemas.microsoft.com/office/spreadsheetml/2022/featurepropertybag">'
    '<bag type="Checkbox"/>'
    '<bag type="XFControls"><bagId k="CellControl">0</bagId></bag>'
    '<bag type="XFComplement"><bagId k="XFControls">1</bagId></bag>'
    '<bag type="XFComplements" extRef="XFComplementsMapperExtRef">'
    '<a k="MappedFeaturePropertyBags"><bagId>2</bagId></a></bag>'
    '</FeaturePropertyBags>'
)
FPB_PART = "xl/featurePropertyBag/featurePropertyBag.xml"
FPB_CONTENT_TYPE = "application/vnd.ms-excel.featurepropertybag+xml"
FPB_REL_TYPE = "http://schemas.microsoft.com/office/2022/11/relationships/FeaturePropertyBag"
XF_EXT = (
    '<extLst><ext uri="{C7286773-470A-42A8-94C5-96B5CB345126}"'
    ' xmlns:xfpb="http://schemas.microsoft.com/office/spreadsheetml/2022/featurepropertybag">'
    '<xfpb:xfComplement i="0"/></ext></extLst>'
)


def patch_chart_xml(xml):
    """Rewrite openpyxl's chart XML with full visible axes (openpyxl's bare
    axis elements never show labels in Excel; these mirror TDM graph.xlsx)."""
    # openpyxl writes the chart with a default (unprefixed) chart namespace and
    # declares xmlns:a inline everywhere; declare it once on the root so the
    # injected fragments can use a: freely.
    xml = xml.replace(
        f'<chartSpace xmlns="{C_NS}">',
        f'<chartSpace xmlns="{C_NS}" xmlns:a="{A_NS}">', 1)

    # Put the title ABOVE the plot area: without <overlay val="0"/> Excel
    # floats it over the plot instead of reserving space for it.
    if "<overlay" not in xml:
        xml = xml.replace("</title>", '<overlay val="0"/></title>', 1)
    if "<autoTitleDeleted" not in xml:
        xml = xml.replace("</title>", '</title><autoTitleDeleted val="0"/>', 1)

    m = re.search(r"<catAx>.*?</catAx>", xml, re.S)
    axid = re.search(r'<axId val="(\d+)"/>', m.group(0)).group(1)
    cross = re.search(r'<crossAx val="(\d+)"/>', m.group(0)).group(1)
    xml = xml.replace(m.group(0), CAT_AX.format(axid=axid, cross=cross))

    m = re.search(r"<valAx>.*?</valAx>", xml, re.S)
    axid = re.search(r'<axId val="(\d+)"/>', m.group(0)).group(1)
    cross = re.search(r'<crossAx val="(\d+)"/>', m.group(0)).group(1)
    minmax = "".join(re.findall(r'<max val="[^"]*"/>|<min val="[^"]*"/>', m.group(0)))
    xml = xml.replace(m.group(0), VAL_AX.format(axid=axid, cross=cross, minmax=minmax))
    return xml


def restore_checkboxes(parts):
    """Re-inject the native checkbox feature openpyxl stripped, and point the
    tick cells (column A booleans on the vital sheets) at a checkbox XF."""
    # Map vital sheet names -> worksheet part names.
    wbxml = parts["xl/workbook.xml"].decode("utf-8")
    rels = parts["xl/_rels/workbook.xml.rels"].decode("utf-8")
    # Parse attribute-by-attribute: openpyxl and Excel order attributes
    # differently, and targets may be absolute (/xl/...) or relative.
    rid_to_target = {}
    for reltag in re.findall(r"<Relationship\b[^>]*/>", rels):
        rid = re.search(r'Id="([^"]+)"', reltag)
        tgt = re.search(r'Target="([^"]+)"', reltag)
        if rid and tgt:
            rid_to_target[rid.group(1)] = tgt.group(1)
    name_to_part = {}
    for sheettag in re.findall(r"<sheet\b[^>]*/>", wbxml):
        name = re.search(r'name="([^"]+)"', sheettag)
        rid = re.search(r'r:id="([^"]+)"', sheettag)
        if not (name and rid):
            continue
        target = rid_to_target.get(rid.group(1), "").lstrip("/")
        if target and not target.startswith("xl/"):
            target = "xl/" + target
        name_to_part[name.group(1)] = target

    # New XF: clone the style the tick cells currently use, add the checkbox ext.
    styles = parts["xl/styles.xml"].decode("utf-8")
    cellxfs = re.search(r"<cellXfs count=\"(\d+)\">(.*)</cellXfs>", styles, re.S)
    xf_count = int(cellxfs.group(1))
    xf_blocks = re.findall(r"<xf\b[^>]*/>|<xf\b[^>]*>.*?</xf>", cellxfs.group(2))

    # Find the style index the tick cells use (first boolean cell in column A).
    tick_re = re.compile(r'(<c r="A(\d+)" s=")(\d+)(" t="b">)')
    sample_s = None
    for sheet in VITAL_SHEETS:
        part = name_to_part.get(sheet)
        if not part or part not in parts:
            continue
        m = tick_re.search(parts[part].decode("utf-8"))
        if m and int(m.group(2)) >= FIRST_DATA_ROW:
            sample_s = int(m.group(3))
            break
    if sample_s is None:
        return parts  # nothing to do

    base_xf = xf_blocks[sample_s]
    if base_xf.endswith("/>"):
        new_xf = base_xf[:-2] + ">" + XF_EXT + "</xf>"
    else:
        new_xf = base_xf.replace("</xf>", XF_EXT + "</xf>")
    new_idx = xf_count
    styles = styles.replace(
        f'<cellXfs count="{xf_count}">',
        f'<cellXfs count="{xf_count + 1}">', 1)
    styles = styles.replace("</cellXfs>", new_xf + "</cellXfs>", 1)
    parts["xl/styles.xml"] = styles.encode("utf-8")

    # Point every column-A boolean tick cell at the checkbox XF.
    for sheet in VITAL_SHEETS:
        part = name_to_part.get(sheet)
        if not part or part not in parts:
            continue
        xml = parts[part].decode("utf-8")
        xml = tick_re.sub(
            lambda m: (m.group(1) + str(new_idx) + m.group(4))
            if int(m.group(2)) >= FIRST_DATA_ROW else m.group(0),
            xml)
        parts[part] = xml.encode("utf-8")

    # Feature property bag part + content type + workbook relationship.
    parts[FPB_PART] = FPB_XML.encode("utf-8")

    ct = parts["[Content_Types].xml"].decode("utf-8")
    if FPB_PART not in ct:
        ct = ct.replace(
            "</Types>",
            f'<Override PartName="/{FPB_PART}" ContentType="{FPB_CONTENT_TYPE}"/></Types>')
        parts["[Content_Types].xml"] = ct.encode("utf-8")

    if FPB_REL_TYPE not in rels:
        max_rid = max(int(n) for n in re.findall(r'Id="rId(\d+)"', rels))
        rels = rels.replace(
            "</Relationships>",
            f'<Relationship Id="rId{max_rid + 1}" Type="{FPB_REL_TYPE}"'
            f' Target="featurePropertyBag/featurePropertyBag.xml"/></Relationships>')
        parts["xl/_rels/workbook.xml.rels"] = rels.encode("utf-8")

    return parts


def patch_package(wb_path):
    """Rewrite the saved .xlsx: chart XML fixes + checkbox restoration."""
    with zipfile.ZipFile(wb_path) as z:
        parts = {info.filename: z.read(info.filename) for info in z.infolist()}

    for name in list(parts):
        if re.fullmatch(r"xl/charts/chart\d+\.xml", name):
            parts[name] = patch_chart_xml(parts[name].decode("utf-8")).encode("utf-8")

    parts = restore_checkboxes(parts)

    with zipfile.ZipFile(wb_path, "w", zipfile.ZIP_DEFLATED) as z:
        for name, data in parts.items():
            z.writestr(name, data)


def main():
    wb_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_WB
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = wb_path.replace(".xlsx", f"_backup_before_skill2_graphs_{stamp}.xlsx")
    shutil.copyfile(wb_path, backup)
    print(f"Backup saved: {backup}")

    wb = load_workbook(wb_path)
    for name in LEGACY_GRAPH_SHEETS:
        if name in wb.sheetnames:
            del wb[name]

    for cfg in VITALS:
        if cfg["sheet"] not in wb.sheetnames:
            print(f"  ! sheet '{cfg['sheet']}' missing; skipping")
            continue
        n = build_vital(wb, wb[cfg["sheet"]], cfg)
        print(f"  {cfg['sheet']}: {n} ticked row(s) plotted")

    wb.save(wb_path)

    patch_package(wb_path)
    print(f"Saved: {wb_path} (axes + checkboxes patched)")


if __name__ == "__main__":
    main()
